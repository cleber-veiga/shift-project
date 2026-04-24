"""
Processador do no de entrada Excel (.xlsx / .xls).

O DuckDB nao lê arquivos Excel nativamente, portanto este no usa ``openpyxl``
para iterar pelas linhas da planilha em modo ``read_only`` (streaming),
escrevendo cada linha no ``JsonlStreamer`` sem nunca carregar a planilha
inteira em RAM.

Fluxo:
  1. Download do arquivo para um Path temporario via ``httpx`` (streaming).
  2. ``openpyxl.load_workbook(read_only=True)`` — uma linha por vez.
  3. Cada linha e serializada em JSON e escrita no JSONL via ``JsonlStreamer``.
  4. Ao finalizar, o DuckDB lê o JSONL e cria a tabela em uma unica passagem.

Configuracao:
- url         : caminho local ou URL HTTP/HTTPS do arquivo (obrigatorio)
- sheet_name  : nome ou indice (0-based) da aba; None = primeira aba (padrao: None)
- header_row  : indice (0-based) da linha de cabecalho (padrao: 0)
- skip_empty  : ignora linhas onde todas as celulas sao None/vazias (padrao: true)
- output_field: nome do campo de saida (padrao: "data")
"""

import tempfile
from pathlib import Path
from typing import Any
from uuid import uuid4

import httpx

from app.core.config import settings
from app.data_pipelines.duckdb_storage import JsonlStreamer
from app.services.workflow.nodes import BaseNodeProcessor, register_processor
from app.services.workflow.nodes.exceptions import NodeProcessingError

_REMOTE_PREFIXES = ("http://", "https://")
_CONNECT_TIMEOUT = 30.0   # segundos para estabelecer conexao
_READ_TIMEOUT = 300.0      # segundos para leitura (arquivos grandes)
_CHUNK_SIZE = 1 << 20      # 1 MB por chunk no download


@register_processor("excel_input")
class ExcelInputNodeProcessor(BaseNodeProcessor):
    """Le planilha Excel e materializa em DuckDB via streaming JSONL."""

    def process(
        self,
        node_id: str,
        config: dict[str, Any],
        context: dict[str, Any],
    ) -> dict[str, Any]:
        resolved_config = self.resolve_data(config, context)
        url = resolved_config.get("url")
        sheet_name = resolved_config.get("sheet_name")  # None = primeira aba
        header_row = int(resolved_config.get("header_row", 0))
        skip_empty = bool(resolved_config.get("skip_empty", True))
        output_field = str(resolved_config.get("output_field", "data"))
        preview_max_rows: int | None = context.get("_preview_max_rows")
        configured_max_rows = resolved_config.get("max_rows")
        max_rows: int | None = (
            preview_max_rows
            if preview_max_rows is not None
            else (int(configured_max_rows) if configured_max_rows is not None else settings.EXTRACT_DEFAULT_MAX_ROWS)
        )

        if not url:
            raise NodeProcessingError(
                f"No excel_input '{node_id}': 'url' e obrigatorio."
            )
        if header_row < 0:
            raise NodeProcessingError(
                f"No excel_input '{node_id}': 'header_row' deve ser >= 0."
            )

        execution_id = str(
            context.get("execution_id") or context.get("workflow_id") or uuid4()
        )

        # Obtem o caminho do arquivo (download se for URL remota)
        local_path, downloaded = _resolve_local_path(node_id, str(url))
        try:
            reference = _stream_excel_to_duckdb(
                local_path=local_path,
                node_id=node_id,
                execution_id=execution_id,
                sheet_name=sheet_name,
                header_row=header_row,
                skip_empty=skip_empty,
                max_rows=max_rows,
            )
        finally:
            if downloaded:
                local_path.unlink(missing_ok=True)

        if reference is None:
            raise NodeProcessingError(
                f"No excel_input '{node_id}': planilha nao continha linhas de dados."
            )

        return {
            "node_id": node_id,
            "status": "completed",
            "row_count": reference.get("_row_count", 0),
            "output_field": output_field,
            output_field: {k: v for k, v in reference.items() if k != "_row_count"},
        }


# ---------------------------------------------------------------------------
# Helpers privados
# ---------------------------------------------------------------------------

def _resolve_local_path(node_id: str, url: str) -> tuple[Path, bool]:
    """
    Retorna (caminho_local, foi_baixado).

    Se for URL remota, faz download em streaming para um arquivo temporario.
    Se for caminho local, retorna o Path diretamente.
    """
    if any(url.lower().startswith(prefix) for prefix in _REMOTE_PREFIXES):
        suffix = ".xlsx" if url.lower().endswith(".xlsx") else ".xls"
        tmp = tempfile.NamedTemporaryFile(suffix=suffix, delete=False)
        tmp_path = Path(tmp.name)
        tmp.close()

        try:
            with httpx.Client(
                timeout=httpx.Timeout(
                    connect=_CONNECT_TIMEOUT, read=_READ_TIMEOUT, write=30.0, pool=5.0
                ),
                follow_redirects=True,
            ) as client:
                with client.stream("GET", url) as response:
                    response.raise_for_status()
                    with tmp_path.open("wb") as fh:
                        for chunk in response.iter_bytes(chunk_size=_CHUNK_SIZE):
                            fh.write(chunk)
        except httpx.HTTPStatusError as exc:
            tmp_path.unlink(missing_ok=True)
            raise NodeProcessingError(
                f"No excel_input '{node_id}': download falhou com HTTP "
                f"{exc.response.status_code} — {url}"
            ) from exc
        except Exception as exc:
            tmp_path.unlink(missing_ok=True)
            raise NodeProcessingError(
                f"No excel_input '{node_id}': erro ao baixar arquivo — {exc}"
            ) from exc

        return tmp_path, True

    # Arquivo local
    local = Path(url)
    if not local.exists():
        raise NodeProcessingError(
            f"No excel_input '{node_id}': arquivo nao encontrado em '{url}'."
        )
    return local, False


def _stream_excel_to_duckdb(
    local_path: Path,
    node_id: str,
    execution_id: str,
    sheet_name: Any,
    header_row: int,
    skip_empty: bool,
    max_rows: int | None = None,
) -> dict[str, Any] | None:
    """
    Itera pela planilha linha a linha (read_only) e materializa via JsonlStreamer.

    Retorna a DuckDbReference + _row_count, ou None se nenhuma linha de dados.
    """
    try:
        import openpyxl  # noqa: PLC0415
    except ImportError as exc:
        raise NodeProcessingError(
            f"No excel_input '{node_id}': 'openpyxl' nao esta instalado. "
            "Instale com: pip install openpyxl"
        ) from exc

    try:
        wb = openpyxl.load_workbook(
            str(local_path), read_only=True, data_only=True
        )
    except Exception as exc:
        raise NodeProcessingError(
            f"No excel_input '{node_id}': nao foi possivel abrir o arquivo Excel — {exc}"
        ) from exc

    try:
        # Seleciona a aba
        if sheet_name is None:
            ws = wb.active
        elif isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name]
        else:
            ws = wb[str(sheet_name)]
    except (KeyError, IndexError) as exc:
        wb.close()
        raise NodeProcessingError(
            f"No excel_input '{node_id}': aba '{sheet_name}' nao encontrada. "
            f"Abas disponiveis: {wb.sheetnames}"
        ) from exc

    try:
        with JsonlStreamer(execution_id, node_id) as streamer:
            headers: list[str] | None = None

            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx < header_row:
                    # Linhas antes do cabecalho sao ignoradas
                    continue

                if row_idx == header_row:
                    # Esta linha e o cabecalho
                    headers = [
                        str(cell).strip() if cell is not None else f"col_{col_idx}"
                        for col_idx, cell in enumerate(row)
                    ]
                    continue

                # Linha de dados
                if headers is None:
                    # Sem cabecalho definido: gera nomes genericos
                    headers = [f"col_{i}" for i in range(len(row))]

                row_dict = {
                    headers[i]: cell
                    for i, cell in enumerate(row)
                    if i < len(headers)
                }

                # Pula linhas completamente vazias se skip_empty=True
                if skip_empty and all(v is None or v == "" for v in row_dict.values()):
                    continue

                # Converte tipos nao serializaveis (datetime, date, etc.)
                row_serializable = {
                    k: _serialize_cell(v) for k, v in row_dict.items()
                }
                streamer.write_row(row_serializable)

                if max_rows is not None and streamer.row_count >= max_rows:
                    break
    finally:
        wb.close()

    if streamer.reference is None:
        return None

    return {**streamer.reference, "_row_count": streamer.row_count}


def _serialize_cell(value: Any) -> Any:
    """Converte tipos do openpyxl em valores serializaveis em JSON."""
    if value is None:
        return None
    # datetime, date, time → ISO string
    if hasattr(value, "isoformat"):
        return value.isoformat()
    # Numeros e strings passam direto
    if isinstance(value, (int, float, str, bool)):
        return value
    # Fallback para string
    return str(value)
