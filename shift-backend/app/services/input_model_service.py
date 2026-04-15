"""
Servico de modelos de entrada (Input Models).

CRUD, geracao de templates e validacao de arquivos uploadados.
"""

from __future__ import annotations

import csv
import io
from typing import Any
from uuid import UUID

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.input_model import InputModel
from app.schemas.input_model import (
    InputModelCreate,
    InputModelSchema,
    InputModelUpdate,
    ValidationResult,
)


class InputModelService:
    """CRUD + template generation + file validation para modelos de entrada."""

    # ─── CRUD ─────────────────────────────────────────────────────────────────

    async def create(
        self,
        db: AsyncSession,
        workspace_id: UUID,
        data: InputModelCreate,
        created_by_id: UUID,
    ) -> InputModel:
        model = InputModel(
            workspace_id=workspace_id,
            name=data.name,
            description=data.description,
            file_type=data.file_type.value,
            schema_def=data.schema_def.model_dump(),
            created_by_id=created_by_id,
        )
        db.add(model)
        await db.flush()
        await db.refresh(model)
        return model

    async def list_by_workspace(
        self, db: AsyncSession, workspace_id: UUID
    ) -> list[InputModel]:
        result = await db.execute(
            select(InputModel)
            .where(InputModel.workspace_id == workspace_id)
            .order_by(InputModel.name)
        )
        return list(result.scalars().all())

    async def get(self, db: AsyncSession, input_model_id: UUID) -> InputModel | None:
        return await db.get(InputModel, input_model_id)

    async def update(
        self,
        db: AsyncSession,
        input_model_id: UUID,
        data: InputModelUpdate,
    ) -> InputModel | None:
        model = await self.get(db, input_model_id)
        if model is None:
            return None

        updates = data.model_dump(exclude_unset=True)
        if "schema_def" in updates and updates["schema_def"] is not None:
            updates["schema_def"] = data.schema_def.model_dump()  # type: ignore[union-attr]
        if "file_type" in updates and updates["file_type"] is not None:
            updates["file_type"] = updates["file_type"].value

        for key, value in updates.items():
            setattr(model, key, value)

        await db.flush()
        await db.refresh(model)
        return model

    async def delete(self, db: AsyncSession, input_model_id: UUID) -> bool:
        model = await self.get(db, input_model_id)
        if model is None:
            return False
        await db.delete(model)
        await db.flush()
        return True

    # ─── Template generation ──────────────────────────────────────────────────

    def generate_template(
        self, model: InputModel
    ) -> tuple[bytes, str, str]:
        """
        Gera um arquivo template vazio a partir do schema.

        Returns: (bytes, filename, content_type)
        """
        schema = InputModelSchema.model_validate(model.schema_def)

        if model.file_type == "data":
            raise ValueError("Modelos do tipo Dados nao possuem template para download.")
        if model.file_type == "csv":
            return self._generate_csv_template(model.name, schema)
        return self._generate_excel_template(model.name, schema)

    def _generate_csv_template(
        self, name: str, schema: InputModelSchema
    ) -> tuple[bytes, str, str]:
        sheet = schema.sheets[0]
        buf = io.StringIO()
        writer = csv.writer(buf, delimiter=";")
        writer.writerow([col.name for col in sheet.columns])
        content = buf.getvalue().encode("utf-8-sig")  # BOM for Excel compat
        return content, f"{name}.csv", "text/csv; charset=utf-8"

    def _generate_excel_template(
        self, name: str, schema: InputModelSchema
    ) -> tuple[bytes, str, str]:
        import openpyxl  # noqa: PLC0415
        from openpyxl.styles import Font, PatternFill  # noqa: PLC0415

        wb = openpyxl.Workbook()
        # Remove default sheet
        if wb.active:
            wb.remove(wb.active)

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        for sheet_def in schema.sheets:
            ws = wb.create_sheet(title=sheet_def.name)
            for col_idx, col_def in enumerate(sheet_def.columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_def.name)
                cell.font = header_font
                cell.fill = header_fill
                # Auto-width based on column name
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(col_idx)
                ].width = max(len(col_def.name) + 4, 12)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return buf.getvalue(), f"{name}.xlsx", content_type

    # ─── File validation ──────────────────────────────────────────────────────

    def validate_file(
        self, model: InputModel, file_bytes: bytes, filename: str
    ) -> ValidationResult:
        """Valida um arquivo uploadado contra o schema do modelo."""
        schema = InputModelSchema.model_validate(model.schema_def)

        if model.file_type == "data":
            from app.schemas.input_model import ValidationResult  # noqa: PLC0415
            return ValidationResult(valid=False, errors=["Modelos do tipo Dados nao aceitam upload de arquivos."])
        if model.file_type == "csv":
            return self._validate_csv(schema, file_bytes, filename)
        return self._validate_excel(schema, file_bytes, filename)

    def _validate_csv(
        self, schema: InputModelSchema, file_bytes: bytes, filename: str
    ) -> ValidationResult:
        errors: list[str] = []

        if not filename.lower().endswith(".csv"):
            errors.append(f"Arquivo '{filename}' nao e um CSV.")
            return ValidationResult(valid=False, errors=errors)

        sheet = schema.sheets[0]
        expected = {col.name.lower() for col in sheet.columns}
        required = {col.name.lower() for col in sheet.columns if col.required}

        try:
            text = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                text = file_bytes.decode("latin-1")
            except UnicodeDecodeError:
                errors.append("Nao foi possivel decodificar o arquivo CSV.")
                return ValidationResult(valid=False, errors=errors)

        reader = csv.reader(io.StringIO(text), delimiter=";")
        header_row = next(reader, None)
        if not header_row:
            errors.append("Arquivo CSV vazio ou sem cabecalho.")
            return ValidationResult(valid=False, errors=errors)

        actual = {h.strip().lower() for h in header_row}

        missing = required - actual
        if missing:
            errors.append(
                f"Colunas obrigatorias ausentes: {', '.join(sorted(missing))}"
            )

        unknown = actual - expected
        if unknown:
            errors.append(
                f"Colunas nao reconhecidas: {', '.join(sorted(unknown))}"
            )

        return ValidationResult(valid=len(errors) == 0, errors=errors)

    def _validate_excel(
        self, schema: InputModelSchema, file_bytes: bytes, filename: str
    ) -> ValidationResult:
        import openpyxl  # noqa: PLC0415

        errors: list[str] = []

        if not filename.lower().endswith((".xlsx", ".xls")):
            errors.append(f"Arquivo '{filename}' nao e um Excel.")
            return ValidationResult(valid=False, errors=errors)

        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(file_bytes), read_only=True, data_only=True
            )
        except Exception:
            errors.append("Nao foi possivel abrir o arquivo Excel.")
            return ValidationResult(valid=False, errors=errors)

        actual_sheets = {s.lower(): s for s in wb.sheetnames}

        for sheet_def in schema.sheets:
            sheet_key = sheet_def.name.lower()
            if sheet_key not in actual_sheets:
                errors.append(f"Aba '{sheet_def.name}' nao encontrada no arquivo.")
                continue

            ws = wb[actual_sheets[sheet_key]]
            header_row = [
                str(cell.value).strip().lower() if cell.value else ""
                for cell in next(ws.iter_rows(min_row=1, max_row=1))
            ]

            expected = {col.name.lower() for col in sheet_def.columns}
            required = {col.name.lower() for col in sheet_def.columns if col.required}
            actual = {h for h in header_row if h}

            missing = required - actual
            if missing:
                errors.append(
                    f"Aba '{sheet_def.name}': colunas obrigatorias ausentes: "
                    f"{', '.join(sorted(missing))}"
                )

            unknown = actual - expected
            if unknown:
                errors.append(
                    f"Aba '{sheet_def.name}': colunas nao reconhecidas: "
                    f"{', '.join(sorted(unknown))}"
                )

        wb.close()
        return ValidationResult(valid=len(errors) == 0, errors=errors)


input_model_service = InputModelService()
